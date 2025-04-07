import openpyxl
from openpyxl.styles import PatternFill
from common_functions import redact_pii  # make sure this refers to your actual PII redaction function

def mask_text(text, labels, gliner_model, llm_model):
    redacted = redact_pii(text, labels, gliner_model, llm_model)
    return redacted

def process_excel(file_path, labels, gliner_model, llm_model):
    try:
        # Load the workbook and keep formatting
        workbook = openpyxl.load_workbook(file_path)
        redacted_file_path = file_path.replace(".xlsx", "_redacted.xlsx")

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip():
                        original_text = cell.value
                        redacted_text = mask_text(original_text, labels, gliner_model, llm_model)
                        if redacted_text != original_text:
                            cell.value = redacted_text
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        workbook.save(redacted_file_path)
        return redacted_file_path

    except Exception as e:
        print(f"Error processing Excel file: {e}")
        return None
